"""
Enhanced XLSX Converter with OCR support for embedded images.
Extracts images from Excel spreadsheets and performs OCR while maintaining cell context.
"""

import io
import os
import sys
import tempfile
from typing import Any, BinaryIO, Optional

from markitdown.converters import HtmlConverter
from markitdown import DocumentConverter, DocumentConverterResult, StreamInfo
from markitdown._exceptions import (
    MissingDependencyException,
    MISSING_DEPENDENCY_MESSAGE,
)
from ._ocr_service import LLMVisionOCRService, format_image_reference

# Try loading dependencies
_xlsx_dependency_exc_info = None
try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()


class XlsxConverterWithOCR(DocumentConverter):
    """
    Enhanced XLSX Converter with OCR support for embedded images.
    Extracts images with their cell positions and performs OCR.
    """

    def __init__(self, ocr_service: Optional[LLMVisionOCRService] = None):
        super().__init__()
        self._html_converter = HtmlConverter()
        self.ocr_service = ocr_service

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension == ".xlsx":
            return True

        if mimetype.startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml"
        ):
            return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[1].with_traceback(
                _xlsx_dependency_exc_info[2]
            )  # type: ignore[union-attr]

        # Get OCR service if available (from kwargs or instance)
        ocr_service: Optional[LLMVisionOCRService] = (
            kwargs.get("ocr_service") or self.ocr_service
        )

        # --- extract_only mode: skip OCR, emit image file references ---
        if kwargs.get("extract_only", False):
            image_output_dir = kwargs.get("image_output_dir") or tempfile.mkdtemp(
                prefix="markitdown_ocr_"
            )
            os.makedirs(image_output_dir, exist_ok=True)
            return self._convert_extract_only(file_stream, image_output_dir, **kwargs)

        if ocr_service:
            # Remove ocr_service from kwargs to avoid duplicate argument error
            kwargs_without_ocr = {k: v for k, v in kwargs.items() if k != "ocr_service"}
            return self._convert_with_ocr(
                file_stream, ocr_service, **kwargs_without_ocr
            )
        else:
            return self._convert_standard(file_stream, **kwargs)

    def _convert_standard(
        self, file_stream: BinaryIO, **kwargs: Any
    ) -> DocumentConverterResult:
        """Standard conversion without OCR."""
        file_stream.seek(0)
        sheets = pd.read_excel(file_stream, sheet_name=None, engine="openpyxl")
        md_content = ""

        for sheet_name in sheets:
            md_content += f"## {sheet_name}\n"
            html_content = sheets[sheet_name].to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

        return DocumentConverterResult(markdown=md_content.strip())

    def _convert_with_ocr(
        self, file_stream: BinaryIO, ocr_service: LLMVisionOCRService, **kwargs: Any
    ) -> DocumentConverterResult:
        """Convert XLSX with image OCR."""
        file_stream.seek(0)
        wb = load_workbook(file_stream)

        md_content = ""

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            md_content += f"## {sheet_name}\n\n"

            # Convert sheet data to markdown table
            file_stream.seek(0)
            try:
                df = pd.read_excel(
                    file_stream, sheet_name=sheet_name, engine="openpyxl"
                )
                html_content = df.to_html(index=False)
                md_content += (
                    self._html_converter.convert_string(
                        html_content, **kwargs
                    ).markdown.strip()
                    + "\n\n"
                )
            except Exception:
                # If pandas fails, just skip the table
                pass

            # Extract and OCR images in this sheet
            images_with_ocr = self._extract_and_ocr_sheet_images(sheet, ocr_service)

            if images_with_ocr:
                md_content += "### Images in this sheet:\n\n"
                for img_info in images_with_ocr:
                    ocr_text = img_info["ocr_text"]
                    md_content += f"*[Image OCR]\n{ocr_text}\n[End OCR]*\n\n"

        return DocumentConverterResult(markdown=md_content.strip())

    def _extract_and_ocr_sheet_images(
        self, sheet: Any, ocr_service: LLMVisionOCRService
    ) -> list[dict]:
        """
        Extract and OCR images from an Excel sheet.

        Args:
            sheet: openpyxl worksheet
            ocr_service: OCR service

        Returns:
            List of dicts with 'cell_ref' and 'ocr_text'
        """
        results = []

        try:
            # Check if sheet has images
            if hasattr(sheet, "_images"):
                for img in sheet._images:
                    try:
                        # Get image data
                        if hasattr(img, "_data"):
                            image_data = img._data()
                        elif hasattr(img, "image"):
                            # Some versions store it differently
                            image_data = img.image
                        else:
                            continue

                        # Create image stream
                        image_stream = io.BytesIO(image_data)

                        # Get cell reference
                        cell_ref = "unknown"
                        if hasattr(img, "anchor"):
                            anchor = img.anchor
                            if hasattr(anchor, "_from"):
                                from_cell = anchor._from
                                if hasattr(from_cell, "col") and hasattr(
                                    from_cell, "row"
                                ):
                                    # Convert column number to letter
                                    col_letter = self._column_number_to_letter(
                                        from_cell.col
                                    )
                                    cell_ref = f"{col_letter}{from_cell.row + 1}"

                        # Perform OCR
                        ocr_result = ocr_service.extract_text(image_stream)

                        if ocr_result.text.strip():
                            results.append(
                                {
                                    "cell_ref": cell_ref,
                                    "ocr_text": ocr_result.text.strip(),
                                    "backend": ocr_result.backend_used,
                                }
                            )

                    except Exception:
                        continue

        except Exception:
            pass

        return results

    def _convert_extract_only(
        self, file_stream: BinaryIO, image_output_dir: str, **kwargs: Any
    ) -> DocumentConverterResult:
        """
        Extract-only mode: extract table data and save embedded images to disk.
        No OCR is performed; images are referenced via file paths.
        """
        from PIL import Image

        file_stream.seek(0)
        wb = load_workbook(file_stream)

        # Filter out ocr_service / extract_only from kwargs passed to HTML converter
        html_kwargs = {
            k: v
            for k, v in kwargs.items()
            if k not in ("ocr_service", "extract_only", "image_output_dir")
        }

        md_content = ""

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            md_content += f"## {sheet_name}\n\n"

            # Convert sheet data to markdown table
            file_stream.seek(0)
            try:
                df = pd.read_excel(
                    file_stream, sheet_name=sheet_name, engine="openpyxl"
                )
                html_content = df.to_html(index=False)
                md_content += (
                    self._html_converter.convert_string(
                        html_content, **html_kwargs
                    ).markdown.strip()
                    + "\n\n"
                )
            except Exception:
                pass

            # Extract images and save to disk
            img_idx = 0
            if hasattr(sheet, "_images"):
                for img in sheet._images:
                    try:
                        # Get image data
                        if hasattr(img, "_data"):
                            image_data = img._data()
                        elif hasattr(img, "image"):
                            image_data = img.image
                        else:
                            continue

                        # Determine extension and dimensions
                        ext = "png"
                        width, height = None, None
                        try:
                            pil_img = Image.open(io.BytesIO(image_data))
                            fmt = pil_img.format
                            if fmt:
                                ext = fmt.lower()
                                if ext == "jpeg":
                                    ext = "jpg"
                            width, height = pil_img.size
                        except Exception:
                            pass

                        # Get cell reference for naming
                        cell_ref = "unknown"
                        if hasattr(img, "anchor"):
                            anchor = img.anchor
                            if hasattr(anchor, "_from"):
                                from_cell = anchor._from
                                if hasattr(from_cell, "col") and hasattr(
                                    from_cell, "row"
                                ):
                                    col_letter = self._column_number_to_letter(
                                        from_cell.col
                                    )
                                    cell_ref = f"{col_letter}{from_cell.row + 1}"

                        filename = f"xlsx_{sheet_name}_{img_idx}.{ext}"
                        # Sanitize filename (sheet names may have spaces/special chars)
                        filename = filename.replace(" ", "_").replace("/", "_")
                        filepath = os.path.join(image_output_dir, filename)
                        with open(filepath, "wb") as f:
                            f.write(image_data)

                        img_ref = format_image_reference(
                            filepath,
                            width=width,
                            height=height,
                            size_bytes=len(image_data),
                        )
                        md_content += f"\n{img_ref}\n\n"
                        img_idx += 1
                    except Exception:
                        continue

        return DocumentConverterResult(markdown=md_content.strip())

    @staticmethod
    def _column_number_to_letter(n: int) -> str:
        """Convert column number to Excel column letter (0-indexed)."""
        result = ""
        n = n + 1  # Make 1-indexed
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result
